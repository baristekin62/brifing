"""ÖZEN Günlük Brifing Uygulaması

D:\\BRIFING içindeki config.json'u okuyarak:
  1. ERP'ye bağlanır ve queries\\ klasöründeki her .sql dosyasını çalıştırır.
  2. Sonuçları HTML rapora dönüştürüp output\\ klasörüne kaydeder.
  3. İsteğe bağlı olarak SMTP profili ile e-posta gönderir.

Kullanım:
  python run_briefing.py            # raporu üretir + kaydeder + e-posta gönderir
  python run_briefing.py --no-email # raporu üretir + kaydeder, e-posta göndermez
  python run_briefing.py --test-smtp# SMTP bağlantısını test eder, sorgu çalıştırmaz
"""

import argparse
import base64
import html
import json
import logging
import os
import smtplib
import sys
import time
from datetime import datetime
from decimal import Decimal
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.utils import formataddr
from pathlib import Path

import pyodbc

BASE_DIR = Path(__file__).resolve().parent
CONFIG_FILE = BASE_DIR / "config.json"
MAIL_LOG_FILE = BASE_DIR / "mail_logs.json"

LOG_DIR = BASE_DIR / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / "briefing.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("ozen.briefing")


# ─── Yardımcılar ─────────────────────────────────────────────────────────


def _load_env() -> dict:
    """Proje kökündeki .env dosyasını okur (K=V satırları)."""
    env = {}
    p = BASE_DIR / ".env"
    if p.exists():
        try:
            for line in p.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
        except Exception:
            pass
    return env


def _decode_secret(value: str) -> str:
    """Şifre şemasını çözer: 'ENV:<VAR>' .env'den, 'ENC:<base64>' veya düz metin."""
    if not value:
        return ""
    if isinstance(value, str) and value.startswith("ENV:"):
        name = value[4:].strip()
        return _load_env().get(name, os.environ.get(name, ""))
    if isinstance(value, str) and value.startswith("ENC:"):
        try:
            return base64.b64decode(value[4:].encode("utf-8")).decode("utf-8")
        except Exception:
            return value
    return value


def load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_mail_log() -> list:
    if not MAIL_LOG_FILE.exists():
        return []
    try:
        return json.loads(MAIL_LOG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def append_mail_log(entry: dict) -> None:
    """E-posta gönderim kaydını mail_logs.json'a ekler (en son 500 kayıt tutulur)."""
    entries = load_mail_log()
    entries.append(entry)
    entries = entries[-500:]
    try:
        MAIL_LOG_FILE.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        logger.warning("Mail logu yazılamadı: %s", e)


def _resolve_path(base_dir: Path, value: str) -> Path:
    p = Path(value)
    if not p.is_absolute():
        p = base_dir / p
    return p


# ─── ERP Katmanı ──────────────────────────────────────────────────────────


def build_erp_conn_str(profile: dict) -> str:
    server = profile.get("server", "")
    port = profile.get("port") or 1433
    database = profile.get("database", "")
    username = profile.get("username", "")
    password = _decode_secret(profile.get("password_encrypted", ""))
    driver = profile.get("driver", "ODBC Driver 17 for SQL Server")
    encrypt = "yes" if profile.get("encrypt") else "no"
    trust = "yes" if profile.get("trust_server_certificate") else "no"
    app_name = profile.get("application_name", "OZEN_Briefing")

    return (
        f"DRIVER={{{driver}}};"
        f"SERVER={server},{port};"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
        f"Encrypt={encrypt};"
        f"TrustServerCertificate={trust};"
        f"ApplicationName={app_name};"
        f"Connection Timeout=15;"
    )


def get_active_erp_profile(config: dict) -> dict:
    source = _resolve_path(BASE_DIR, config.get("erp_connection_source", ""))
    data = load_json(source)
    profiles = data.get("profiles", [])
    active_id = data.get("active_profile_id")
    for p in profiles:
        if p.get("profile_id") == active_id and p.get("is_active", True):
            return p
    for p in profiles:
        if p.get("is_active", True):
            return p
    if profiles:
        return profiles[0]
    raise RuntimeError(f"ERP bağlantı profili bulunamadı: {source}")


def run_sql_query(conn: pyodbc.Connection, sql: str, timeout_sec: int = 120) -> dict:
    """Tek SQL sorgusu çalıştırır; sonuç sözlük olarak döner."""
    t0 = time.perf_counter()
    cur = conn.cursor()
    cur.execute(sql)
    columns = [col[0] for col in cur.description] if cur.description else []
    rows = []
    if columns:
        try:
            rows = cur.fetchall()
        except Exception:
            rows = []
    duration_ms = round((time.perf_counter() - t0) * 1000, 1)
    result = {
        "success": True,
        "columns": columns,
        "rows": [tuple(r) for r in rows],
        "row_count": len(rows),
        "duration_ms": duration_ms,
    }
    cur.close()
    return result


# ─── SMTP Katmanı ─────────────────────────────────────────────────────────


def get_smtp_profile(config: dict) -> dict:
    source = _resolve_path(BASE_DIR, config.get("smtp_profiles_source", ""))
    data = load_json(source)
    profiles = data.get("profiles", [])
    profile_id = config.get("smtp_profile_id")
    for p in profiles:
        if p.get("id") == profile_id:
            return p
    for p in profiles:
        if p.get("is_default"):
            return p
    if profiles:
        return profiles[0]
    raise RuntimeError(f"SMTP profili bulunamadı: {source}")


def smtp_connect(profile: dict) -> smtplib.SMTP:
    host = profile.get("host")
    port = int(profile.get("port", 587))
    timeout = int(profile.get("timeout_sec", 15))
    use_ssl = bool(profile.get("use_ssl", False))

    if use_ssl:
        server = smtplib.SMTP_SSL(host, port, timeout=timeout)
    else:
        server = smtplib.SMTP(host, port, timeout=timeout)
        server.ehlo()
        if profile.get("use_tls"):
            server.starttls()
            server.ehlo()

    if profile.get("use_auth") and profile.get("username"):
        server.login(profile.get("username"), _decode_secret(profile.get("password_enc", "")))
    return server


def send_email_via_smtp(profile: dict, to_emails, subject: str, body_html: str,
                        cc_emails=None, bcc_emails=None, attachments=None) -> dict:
    t0 = time.perf_counter()
    cc_emails = cc_emails or []
    bcc_emails = bcc_emails or []
    attachments = attachments or []
    from_addr = profile.get("from_address") or profile.get("username")
    display_name = profile.get("display_name") or "ÖZEN AI Platform"

    msg = MIMEMultipart("mixed")
    msg["From"] = formataddr((display_name, from_addr))
    msg["To"] = ", ".join(to_emails)
    if cc_emails:
        msg["Cc"] = ", ".join(cc_emails)
    if bcc_emails:
        msg["Bcc"] = ", ".join(bcc_emails)
    msg["Subject"] = subject
    if profile.get("reply_to"):
        msg["Reply-To"] = profile.get("reply_to")

    # HTML gövde (alternative katmanı)
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(body_html, "html", "utf-8"))
    msg.attach(alt)

    # Ekler
    for fpath in attachments:
        fpath = Path(fpath)
        if not fpath.exists():
            continue
        import mimetypes
        ctype, _ = mimetypes.guess_type(str(fpath))
        ctype = ctype or "application/octet-stream"
        maintype, subtype = ctype.split("/", 1)
        try:
            with open(fpath, "rb") as f:
                payload = f.read()
            part = MIMEBase(maintype, subtype)
            part.set_payload(payload)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment",
                            filename=("utf-8", "", fpath.name))
            msg.attach(part)
        except Exception as e:
            logger.warning("Ek eklenemedi %s: %s", fpath, e)

    server = None
    try:
        server = smtp_connect(profile)
        all_recipients = list(set(list(to_emails) + cc_emails + bcc_emails))
        server.sendmail(from_addr, all_recipients, msg.as_string())
        dt = round((time.perf_counter() - t0) * 1000, 1)
        return {"success": True, "latency_ms": dt}
    except Exception as e:
        dt = round((time.perf_counter() - t0) * 1000, 1)
        return {"success": False, "error": str(e), "latency_ms": dt}
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass


# ─── Rapor Üretimi ────────────────────────────────────────────────────────


def build_excel_report(query_results: list, output_path: Path) -> Path:
    """Her sorguyu ayrı bir Excel sayfasına yazar.

    - Sayfa adı: sorgu başlığı (maks 31 karakter)
    - İlk satır başlıklar, altında veriler
    - Sayısal sütunlar Türkçe sayı biçiminde yazılır
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # varsayılan boş sayfayı sil

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for q in query_results:
        title = (q.get("title") or "Sorgu")[:31]
        ws = wb.create_sheet(title=title)

        if not q.get("success"):
            ws.cell(row=1, column=1, value="Sorgu hatası: " + str(q.get("error")))
            continue

        columns = q.get("columns", [])
        rows = q.get("rows", [])

        # Başlıklar
        for ci, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=ci, value=str(col))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Veriler
        for ri, row in enumerate(rows, start=2):
            for ci, value in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=_excel_value(value))
                cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float, Decimal)) else "left")

        # Sütun genişliklerini ayarla
        for ci in range(1, len(columns) + 1):
            letter = get_column_letter(ci)
            max_len = 8
            for cell in ws[letter][: min(ws.max_row, 200) + 1]:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 40)

        # Başlık satırını dondur
        ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def _excel_value(value):
    """Excel için değeri hazırlar; Decimal/float'ı 2 ondalıklı sayı yapar."""
    if isinstance(value, (int, float, Decimal)):
        return round(float(value), 2)
    if value is None:
        return ""
    return value


def value_to_str(value) -> str:
    if value is None:
        return "-"
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float, Decimal)):
        # Türkçe format: 18.235,79 (nokta binlik, virgül ondalık)
        if isinstance(value, int):
            return f"{value:,}".replace(",", ".")
        s = f"{float(value):,.2f}"                       # 18,235.79
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")  # 18.235,79
        return s
    return str(value)


def render_table(columns, rows, max_rows: int = 200) -> str:
    show = rows[:max_rows]
    truncated = len(rows) > max_rows

    thead = "".join(f"<th>{html.escape(str(c))}</th>" for c in columns)
    body = ""
    for row in show:
        tds = "".join(f"<td>{html.escape(value_to_str(v))}</td>" for v in row)
        body += f"<tr>{tds}</tr>"

    note = f"<p class='note'>Toplam {len(rows)} satır gösterildi (ilk {max_rows}).</p>" if truncated else ""
    return (
        f"<table><thead><tr>{thead}</tr></thead><tbody>{body}</tbody></table>{note}"
    )


def _resolve_kpi_value(q: dict, kpi: dict):
    """KPI tanımındaki (column,row) değerini sorgu sonucundan çeker."""
    if not q.get("success"):
        return None
    columns = q.get("columns", [])
    rows = q.get("rows", [])
    col_name = str(kpi.get("column", "")).strip().lower()
    col_idx = None
    for i, c in enumerate(columns):
        if str(c).strip().lower() == col_name:
            col_idx = i
            break
    if col_idx is None:
        return None
    row_idx = int(kpi.get("row", 0))
    if row_idx < len(rows):
        return rows[row_idx][col_idx]
    return None


def build_kpi_cards(query_results: list) -> str:
    """Manifest'te tanımlı KPI'ları ve tek satırlık özet sorguların değerlerini kart yapar."""
    cards = []
    for q in query_results:
        if not q.get("success"):
            continue
        # 1) Manifestte tanımlı KPI'lar
        for kpi in q.get("kpis", []):
            label = kpi.get("label") or kpi.get("column")
            if not label:
                continue
            val = _resolve_kpi_value(q, kpi)
            if val is None:
                continue
            cards.append((label, value_to_str(val)))
        # 2) Tek satırlık özet sorgular: her sayısal sütunu otomatik kart yap
        if len(q.get("rows", [])) == 1 and not q.get("kpis"):
            row = q["rows"][0]
            for ci, c in enumerate(q["columns"]):
                if ci < len(row):
                    v = row[ci]
                    if isinstance(v, (int, float, Decimal)) and v != 0:
                        cards.append((str(c), value_to_str(v)))

    if not cards:
        return ""

    items = "".join(
        f"<div class='kpi-card'><span class='kpi-label'>{html.escape(label)}</span>"
        f"<span class='kpi-value'>{html.escape(val)}</span></div>"
        for label, val in cards[:12]
    )
    return f"<div class='kpi-grid'>{items}</div>"


def build_report_html(config: dict, query_results: list, generated_at: str, insights_html: str = "") -> str:
    title = config.get("report_title", "ÖZEN Günlük Yönetici Brifingi")
    date_str = datetime.now().strftime("%d.%m.%Y")

    sections_html = ""
    for q in query_results:
        status_icon = "✅" if q["success"] else "❌"
        duration = q.get("duration_ms", 0)
        badge = f"<span class='badge'>{q['row_count']} satır · {duration:.0f} ms</span>"
        if q["success"]:
            content = render_table(q["columns"], q["rows"])
        else:
            content = f"<p class='error'>Sorgu hatası: {html.escape(str(q['error']))}</p>"
        sections_html += (
            f"<div class='section'>"
            f"<h2>{status_icon} {html.escape(q['title'])} {badge}</h2>"
            f"{content}"
            f"</div>"
        )

    return f"""<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="utf-8">
<title>{html.escape(title)} - {date_str}</title>
<style>
body {{ font-family: 'Segoe UI', Tahoma, sans-serif; background: #f1f5f9; margin: 0; padding: 24px; color: #1e293b; }}
.container {{ max-width: 1500px; margin: 0 auto; }}
.header {{ background: linear-gradient(135deg, #0f172a, #1e3a8a); color: #fff; padding: 24px 32px; border-radius: 12px; margin-bottom: 20px; }}
.header h1 {{ margin: 0; font-size: 22px; }}
.header p {{ margin: 6px 0 0; opacity: .85; font-size: 13px; }}
.kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 14px; margin-bottom: 20px; }}
.kpi-card {{ background: #fff; border-radius: 10px; padding: 16px 20px; box-shadow: 0 1px 3px rgba(0,0,0,.08); border-left: 4px solid #1e3a8a; }}
.kpi-label {{ display: block; font-size: 12px; color: #64748b; text-transform: uppercase; letter-spacing: .5px; margin-bottom: 4px; }}
.kpi-value {{ font-size: 22px; font-weight: 700; color: #0f172a; }}
.insight-box {{ background: #fff; border-radius: 10px; padding: 20px 24px; margin-bottom: 18px; box-shadow: 0 1px 3px rgba(0,0,0,.08); border-left: 4px solid #2563eb; }}
.insight-box h3, .anom-box h3 {{ margin: 0 0 10px; font-size: 15px; color: #0f172a; }}
.insight-item {{ padding: 6px 0; border-bottom: 1px dashed #e2e8f0; font-size: 13.5px; line-height: 1.5; }}
.insight-item:last-child {{ border-bottom: none; }}
.anom-box {{ background: #fffbeb; border: 1px solid #fde68a; border-radius: 10px; padding: 18px 22px; margin-bottom: 18px; }}
.anom-item {{ display: flex; gap: 8px; align-items: flex-start; padding: 5px 0; font-size: 13.5px; }}
.anom-icon {{ font-size: 16px; }} 
.section {{ background: #fff; border-radius: 10px; padding: 20px 24px; margin-bottom: 18px; box-shadow: 0 1px 3px rgba(0,0,0,.08); overflow-x: auto; }}
.section h2 {{ font-size: 16px; margin: 0 0 14px; color: #0f172a; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px; }}
.badge {{ float: right; font-size: 12px; background: #e2e8f0; color: #475569; padding: 3px 10px; border-radius: 12px; font-weight: 600; }}
table {{ width: 100%; border-collapse: collapse; font-size: 13px; min-width: 600px; }}
th {{ background: #0f172a; color: #fff; padding: 8px 10px; text-align: left; position: sticky; top: 0; white-space: nowrap; }}
td {{ padding: 7px 10px; border-bottom: 1px solid #e2e8f0; white-space: nowrap; }}
tr:nth-child(even) {{ background: #f8fafc; }}
.note {{ color: #94a3b8; font-size: 12px; margin: 10px 0 0; }}
.error {{ color: #b91c1c; background: #fef2f2; padding: 12px; border-radius: 6px; }}
.footer {{ text-align: center; color: #94a3b8; font-size: 12px; margin-top: 24px; }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>{html.escape(title)}</h1>
    <p>{date_str} · Rapor üretim zamanı: {generated_at}</p>
  </div>
  {build_kpi_cards(query_results)}
  {insights_html}
  {sections_html}
  <p class="footer">ÖZEN GROSS Günlük Brifing · otomatik üretildi </p>
</div>
</body>
</html>"""


# ─── Ana Akış ─────────────────────────────────────────────────────────────


# ─── Sorgu Yönetimi ───────────────────────────────────────────────────────


def load_query_manifest(config: dict) -> list:
    """queries_meta.json manifestini okur; yoksa queries/*.sql üzerinden üretir.

    Manifest elemanı: {file, title, enabled, order, kpis}
    """
    queries_dir = _resolve_path(BASE_DIR, config.get("queries_dir", "queries"))
    manifest_path = BASE_DIR / "queries_meta.json"
    default_title = lambda f: f.stem.replace("_", " ")

    if manifest_path.exists():
        try:
            data = json.loads(manifest_path.read_text(encoding="utf-8"))
            items = data.get("queries", [])
        except Exception:
            items = []
        if items:
            ordered = sorted(items, key=lambda x: (x.get("order", 999), x.get("file", "")))
            result = []
            for it in ordered:
                fpath = queries_dir / it.get("file", "")
                if fpath.exists():
                    result.append({
                        "file": fpath,
                        "title": it.get("title") or default_title(fpath),
                        "enabled": bool(it.get("enabled", True)),
                        "kpis": it.get("kpis", []),
                        "summary": it.get("summary", {}),
                    })
            return result

    # Manifest yok: tüm .sql dosyaları aktif varsayılır
    sql_files = sorted(queries_dir.glob("*.sql"))
    return [{"file": f, "title": default_title(f), "enabled": True} for f in sql_files]


# ─── Brifing Profilleri ───────────────────────────────────────────────────


def get_profile(config: dict, profile_id: str = None) -> dict:
    """briefing_profiles.json dosyasından profili getirir.

    profile_id verilmezse aktif (enabled) ilk profil veya varsayılan döner.
    Dosya yoksa config'ten tek profil üretir.
    """
    profiles_file = BASE_DIR / "briefing_profiles.json"
    if profiles_file.exists():
        try:
            data = json.loads(profiles_file.read_text(encoding="utf-8"))
            profiles = data.get("profiles", [])
        except Exception:
            profiles = []
        if profiles:
            if profile_id:
                for p in profiles:
                    if p.get("id") == profile_id:
                        return p
                return None
            for p in profiles:
                if p.get("is_default", False) or p.get("enabled", True):
                    return p
            return profiles[0]

    # Dosya yok: config'ten tek profil üret
    return {
        "id": "default",
        "name": config.get("report_title", "Günlük Brifing"),
        "enabled": True,
        "is_default": True,
        "to": config.get("recipients", {}).get("to", []),
        "cc": config.get("recipients", {}).get("cc", []),
        "bcc": config.get("recipients", {}).get("bcc", []),
        "subject_template": config.get("subject_template", "ÖZEN Günlük Brifingi ({date})"),
        "query_files": [],
        "send_time": config.get("send_time", "08:00"),
        "send_email": config.get("send_email", True),
        "save_output": config.get("save_output", True),
    }


def get_profile_subject(profile: dict, config: dict) -> str:
    tpl = profile.get("subject_template") or config.get("subject_template", "ÖZEN Günlük Brifingi ({date})")
    return tpl.format(date=datetime.now().strftime("%d.%m.%Y"))


def run_briefing(config: dict, send_email_flag: bool = True, profile_id: str = None) -> dict:
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    # 0) Profil seçimi
    profile = get_profile(config, profile_id)
    if profile is None:
        return {"success": False, "error": f"Brifing profili bulunamadı: {profile_id}", "output_file": None, "email": None}

    # 1) ERP'ye bağlan
    erp_profile = get_active_erp_profile(config)
    conn_str = build_erp_conn_str(erp_profile)
    logger.info("ERP bağlantısı kuruluyor: %s / %s", erp_profile.get("server"), erp_profile.get("database"))
    conn = pyodbc.connect(conn_str)
    conn.timeout = 120

    # 2) Sorguları çalıştır (profilin kendi listesi)
    query_items = load_query_manifest(config)
    selected_files = set(profile.get("query_files", []))
    if selected_files:
        enabled_items = [it for it in query_items if it["enabled"] and it["file"].name in selected_files]
    else:
        enabled_items = [it for it in query_items if it["enabled"]]
    if not enabled_items:
        logger.warning("Aktif sorgu bulunamadı (profil: %s).", profile.get("id"))
        conn.close()
        return {"success": False, "error": "Aktif sorgu yok", "output_file": None, "email": None}

    query_results = []
    for it in enabled_items:
        f = it["file"]
        logger.info("Sorgu çalıştırılıyor: %s", f.name)
        sql = f.read_text(encoding="utf-8")
        try:
            res = run_sql_query(conn, sql)
            res["title"] = it["title"]
            res["kpis"] = it.get("kpis", [])
            res["summary"] = it.get("summary", {})
            logger.info("  → %s satır, %.0f ms", res["row_count"], res.get("duration_ms", 0))
        except Exception as e:
            res = {
                "success": False,
                "columns": [],
                "rows": [],
                "row_count": 0,
                "duration_ms": 0,
                "error": str(e),
                "title": it["title"],
                "kpis": it.get("kpis", []),
            }
            logger.error("  → hata: %s", e)
        query_results.append(res)

    try:
        conn.close()
    except Exception:
        pass

    # 3) HTML rapor üret ve kaydet
    insights_html = ""
    insights_cfg = config.get("insights", {})
    if insights_cfg.get("enabled", True):
        try:
            import insights as _ins
            insights_html = _ins.build_insights(config, query_results)
            if insights_cfg.get("save_history", True):
                _ins.save_history(query_results)
                logger.info("İçgörü geçmişi kaydedildi (%d bölüm).", len(query_results))
        except Exception as exc:
            logger.warning("İçgörü üretilemedi: %s", exc)
            insights_html = ""
    html_body = build_report_html(config, query_results, generated_at, insights_html=insights_html)
    output_file = None
    excel_file = None
    if config.get("save_output", True):
        output_dir = _resolve_path(BASE_DIR, config.get("output_dir", "output"))
        output_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_file = output_dir / f"brifing_{stamp}.html"
        output_file.write_text(html_body, encoding="utf-8")
        logger.info("Rapor kaydedildi: %s", output_file)
        try:
            excel_file = output_dir / f"brifing_{stamp}.xlsx"
            build_excel_report(query_results, excel_file)
            logger.info("Excel raporu kaydedildi: %s", excel_file)
        except Exception as exc:
            excel_file = None
            logger.error("Excel üretilemedi: %s", exc)

    # 4) E-posta gönder (profilin alıcı ve konu ayarları)
    email_result = None
    prof_send_email = profile.get("send_email", config.get("send_email", True))
    if send_email_flag and prof_send_email:
        smtp_profile = get_smtp_profile(config)
        subject = get_profile_subject(profile, config)
        to_emails = profile.get("to") or config.get("recipients", {}).get("to", [])
        cc_emails = profile.get("cc") or []
        bcc_emails = profile.get("bcc") or []
        logger.info("E-posta gönderiliyor: %s → %s", smtp_profile.get("id"), ", ".join(to_emails))
        email_result = send_email_via_smtp(smtp_profile, to_emails, subject, html_body,
                                           cc_emails=cc_emails, bcc_emails=bcc_emails,
                                           attachments=[excel_file] if excel_file else None)
        if email_result.get("success"):
            logger.info("E-posta gönderildi (%.0f ms)", email_result.get("latency_ms", 0))
        else:
            logger.error("E-posta gönderilemedi: %s", email_result.get("error"))
        append_mail_log({
            "time": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            "profile_id": profile.get("id"),
            "profile_name": profile.get("name"),
            "subject": subject,
            "to": to_emails,
            "cc": cc_emails,
            "bcc": bcc_emails,
            "success": bool(email_result.get("success")),
            "error": email_result.get("error"),
            "latency_ms": email_result.get("latency_ms"),
            "attachment": str(excel_file) if excel_file else None,
            "sections": len(query_results),
        })

    return {
        "success": True,
        "output_file": str(output_file) if output_file else None,
        "excel_file": str(excel_file) if excel_file else None,
        "email": email_result,
        "sections": len(query_results),
    }


def test_smtp(config: dict) -> dict:
    profile = get_smtp_profile(config)
    t0 = time.perf_counter()
    server = None
    try:
        server = smtp_connect(profile)
        dt = round((time.perf_counter() - t0) * 1000, 1)
        return {"success": True, "latency_ms": dt, "message": f"SMTP bağlantısı başarılı ({profile.get('host')}:{profile.get('port')})"}
    except Exception as e:
        dt = round((time.perf_counter() - t0) * 1000, 1)
        return {"success": False, "latency_ms": dt, "error": str(e)}
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass


def main() -> int:
    parser = argparse.ArgumentParser(description="ÖZEN Günlük Brifing Uygulaması")
    parser.add_argument("--no-email", action="store_true", help="E-posta gönderme, sadece rapor üret")
    parser.add_argument("--test-smtp", action="store_true", help="SMTP bağlantısını test et, sorgu çalıştırma")
    parser.add_argument("--profile", type=str, default=None, help="Brifing profili ID'si (briefing_profiles.json)")
    args = parser.parse_args()

    config = load_json(CONFIG_FILE)

    if args.test_smtp:
        result = test_smtp(config)
        if result.get("success"):
            logger.info("✅ %s (%.0f ms)", result["message"], result["latency_ms"])
        else:
            logger.error("❌ SMTP testi başarısız: %s", result.get("error"))
            return 1
        return 0

    result = run_briefing(config, send_email_flag=not args.no_email, profile_id=args.profile)

    if not result.get("success"):
        logger.error("Brifing üretilemedi: %s", result.get("error"))
        return 1

    logger.info("Brifing tamamlandı: %d bölüm, dosya=%s", result.get("sections", 0), result.get("output_file"))

    email = result.get("email")
    if email is not None:
        if email.get("success"):
            logger.info("✅ E-posta gönderildi (%.0f ms)", email.get("latency_ms", 0))
        else:
            logger.error("❌ E-posta hatası: %s", email.get("error"))
            return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
